#-*-coding:utf-8-*-
# @time     :2019/7/10/010 17:20
# @Author   :python13_marimo
# @Email    :2670622242@qq.com
# @File     :do_excel.py
# @Sofeware :PyCharm Community Edition
from openpyxl import load_workbook
class cases:
    def __init__(self,case_id=None,case_title=None,case_method=None,case_url=None,case_data=None,case_expect=None):
        self.case_id=case_id
        self.case_title=case_title
        self.method=case_method
        self.case_url=case_url
        self.case_data=case_data
        self.case_expect=case_expect




class doExcel:
    def __init__(self,file_name):
        self.file_name=file_name
        self.file=load_workbook(self.file_name)
        # self.sheet=self.file[sheet]
    def get_data(self,sheet):

        sheet_1=self.file[sheet]
        datas=[]

        for i in range(2,sheet_1.max_row+1):
            case = cases()
            case.case_id=sheet_1.cell(row=i,column=1).value
            case.case_title=sheet_1.cell(row=i,column=2).value
            case.case_method=sheet_1.cell(row=i,column=3).value
            case.case_url=sheet_1.cell(row=i, column=4).value
            case.case_data=sheet_1.cell(row=i, column=5).value
            case.case_expect=sheet_1.cell(row=i, column=6).value
            datas.append(case)

        return datas
    def write_data(self,sheet,w_row,actual,result):
        sheet_1=self.file[sheet]
        sheet_1.cell(row=w_row,column=7).value=actual
        sheet_1.cell(row=w_row, column=8).value = result
        # self.sheet.cell(row,column).value = value_1

        self.file.save(filename=self.file_name)

if __name__ == '__main__':
    from common.context import str_re
    from common import route
    read = doExcel(route.exc_1)
    cases_2 = read.get_data('audit')
    for case in cases_2:
        case.case_data=str_re(case.case_data)
        print(case.case_data)



